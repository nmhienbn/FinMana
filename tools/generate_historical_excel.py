from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROWS = [
    ("01/06/2026 18:41", -38000, 11622891, "DEN: HO KINH DOANH TAP HOA TIEN DUNG - V3GSTPB00601229 | ND: NGUYEN MINH ...", True),
    ("01/06/2026 19:05", -610869, 11012022, "IND: AP-CASHIN-0364567689-O4CI1808t4efd-bank2wallet", False),
    ("02/06/2026 00:42", -10000, 11002022, "IND: QAGQLZ0388 SEPAY10142 1 SEVQR ZYPAGE 1780335741847", False),
    ("02/06/2026 00:49", -2000, 11000022, "IND: QAGQLZ0388 SEPAY10142 1 SEVQR ZYPAGE 1780336166929", False),
    ("02/06/2026 18:59", -25000, 10975022, "DEN: MOMO - PMC2605619300000080 | ND: NGUYEN MINH HIEN chuyen tien", False),
    ("03/06/2026 19:00", -115000, 10860022, "IND: Q0000147Nk APPMI7197109 1 TT STK 74548397979", False),
    ("04/06/2026 09:00", -20000, 10840022, "DEN: NGUYEN ANH TUAN - MBF2922AF2AB0C5545F | ND: NGUYEN MINH HIEN chuyen tien", False),
    ("04/06/2026 09:10", -460000, 10380022, "IND: NGUYEN MINH HIEN chuyen tien", False),
    ("04/06/2026 13:33", -190000, 10190022, "IND: NGUYEN MINH HIEN chuyen tien", False),
    ("04/06/2026 18:38", -8050000, 2140022, "IND: Chuyen tien cashbycode mobileretail", False),
    ("04/06/2026 18:47", -35000, 2105022, "IND: QACKXT5786 MBMMS01103304 1 VQRLOAMB25051366829", False),
    ("04/06/2026 19:50", 3120000, 5225022, "TU: VU VAN MINH - 104809992004 | ND: VU VAN MINH Chuyen tien - Ma GD ACSP/2070...", True),
    ("05/06/2026 10:18", 16080000, 21305022, "TU: CT TNHH BIOTURING VIET NAM - 10004438 | ND: Bioturing Vietnam Co. Lt...", True),
    ("05/06/2026 19:18", -55000, 21250022, "DEN: -615612595001 | ND: POSRETAIL-615612595001-55000-VND-...", True),
    ("05/06/2026 20:05", 500000, 21750022, "TU: VU VAN MINH - 104809992004 | ND: VU VAN MINH Chuyen tien - Ma GD ACSP/XX947645", False),
    ("05/06/2026 20:42", -500000, 21250022, "DEN: VU VAN MINH - 104809992004 | ND: NGUYEN MINH HIEN chuyen tien", False),
    ("05/06/2026 21:37", -5000, 21245022, "DEN: NGUYEN THANH CHUNG - 0031000298076 | ND: MBCT NGUYEN MINH HIEN chuyen tien...", True),
    ("06/06/2026 08:22", -37000, 21208022, "DEN: -615601186951 | ND: POSRETAIL-615601186951-37000-VND-...", True),
    ("06/06/2026 09:18", -15000, 21193022, "IND: NGUYEN MINH HIEN chuyen tien", False),
    ("06/06/2026 10:07", -20000, 21173022, "DEN: Do Thi Tuyen - 2602205474548 | ND: NGUYEN MINH HIEN chuyen tien", False),
    ("06/06/2026 23:59", 2855000, 24028022, "TU: NGO THANH MINH - 1031821535 | ND: MBVCB.14556645758.252066.NGO THA...", True),
    ("07/06/2026 12:44", -107900, 23920122, "IND: AP-CASHIN-0364567689-O4CI180evm58I-bank2wallet", False),
    ("07/06/2026 15:11", -40000, 23880122, "DEN: -615808221130 | ND: POSRETAIL-615808221130-40000-VND-...", True),
    ("07/06/2026 16:44", -5000, 23875122, "DEN: BUI HUY TUONG - AGBVMS0303160531860 | ND: MBCT MCRJLNBYWSWYHNP 52295094...", True),
    ("07/06/2026 19:32", -62000, 23813122, "DEN: -615812851072 | ND: POSRETAIL-615812851072-62000-VND-...", True),
    ("07/06/2026 22:09", -168800, 23644322, "DEN: MINITI VIET NAM SERVICE AND TRADE COMPANY LIMITED - MD181780844952...", True),
    ("07/06/2026 22:17", -16000, 23628322, "DEN: VU THI NHI - MS00T01791615076842 | ND: MBCT NGUYEN MINH HIEN chuyen tien D2ZU...", True),
    ("08/06/2026 10:47", -66000, 23562322, "DEN: TRAN ANH DUC - 2154758379 | ND: NGUYEN MINH HIEN chuyen tien", False),
    ("08/06/2026 20:02", -19000, 23543322, "IND: QRCODE VNPAY QRCODE 036456768926060820020159369 BP00...", True),
    ("09/06/2026 09:01", -20000, 23523322, "DEN: NGUYEN MINH HIEN/2154758078 - 0364567689 | ND: NGUYEN MINH HIEN chuyen tien", False),
]


def main():
    output = Path("FinMana_Lich_su_01-09_06_2026.xlsx")
    wb = Workbook()
    tx = wb.active
    tx.title = "Giao dịch"
    headers = ["Thời gian", "Loại", "Số tiền", "Số dư sau GD", "Danh mục", "Ghi chú", "Nguồn", "Nội dung", "Cần rà soát"]
    tx.append(headers)

    for raw_time, signed_amount, balance, content, truncated in ROWS:
        occurred = datetime.strptime(raw_time, "%d/%m/%Y %H:%M")
        tx.append([
            occurred,
            "INCOME" if signed_amount > 0 else "EXPENSE",
            abs(signed_amount),
            balance,
            "Chưa phân loại",
            "",
            "MB Bank / Be",
            content,
            "Có - nội dung bị cắt trong ảnh" if truncated else "",
        ])

    summary = wb.create_sheet("Tổng hợp ngày")
    summary.append(["Ngày", "Thu", "Chi", "Dòng tiền ròng", "Số giao dịch"])
    dates = sorted({datetime.strptime(row[0], "%d/%m/%Y %H:%M").date() for row in ROWS})
    for date in dates:
        income = sum(max(row[1], 0) for row in ROWS if datetime.strptime(row[0], "%d/%m/%Y %H:%M").date() == date)
        expense = sum(abs(min(row[1], 0)) for row in ROWS if datetime.strptime(row[0], "%d/%m/%Y %H:%M").date() == date)
        count = sum(1 for row in ROWS if datetime.strptime(row[0], "%d/%m/%Y %H:%M").date() == date)
        summary.append([date, income, expense, income - expense, count])
    summary.append(["TỔNG", sum(max(r[1], 0) for r in ROWS), sum(abs(min(r[1], 0)) for r in ROWS), sum(r[1] for r in ROWS), len(ROWS)])

    notes = wb.create_sheet("Ghi chú")
    notes.append(["Nội dung"])
    notes.append(["Dữ liệu được nhập thủ công từ 8 ảnh chụp lịch sử biến động số dư do người dùng cung cấp."])
    notes.append(["Dòng hóa đơn điện tử rút tiền 8.050.000 VND không được nhập riêng vì trùng giao dịch -8.050.000 VND lúc 18:38 ngày 04/06/2026."])
    notes.append(["Các nội dung có dấu ... đã bị cắt trong ảnh và được đánh dấu Cần rà soát."])
    notes.append(["Số tiền trong cột Số tiền luôn là số dương; chiều giao dịch nằm ở cột Loại."])

    header_fill = PatternFill("solid", fgColor="167D5B")
    for sheet in (tx, summary, notes):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

    tx.column_dimensions["A"].width = 20
    tx.column_dimensions["B"].width = 12
    tx.column_dimensions["C"].width = 16
    tx.column_dimensions["D"].width = 18
    tx.column_dimensions["E"].width = 20
    tx.column_dimensions["F"].width = 24
    tx.column_dimensions["G"].width = 18
    tx.column_dimensions["H"].width = 75
    tx.column_dimensions["I"].width = 30
    for row in tx.iter_rows(min_row=2):
        row[0].number_format = "dd/mm/yyyy hh:mm"
        row[2].number_format = '#,##0 "VND"'
        row[3].number_format = '#,##0 "VND"'
        row[7].alignment = Alignment(wrap_text=True, vertical="top")

    for column, width in {"A": 16, "B": 18, "C": 18, "D": 20, "E": 16}.items():
        summary.column_dimensions[column].width = width
    for row in summary.iter_rows(min_row=2):
        row[0].number_format = "dd/mm/yyyy"
        for cell in row[1:4]:
            cell.number_format = '#,##0 "VND"'
    notes.column_dimensions["A"].width = 120

    wb.save(output)

    check = load_workbook(output, read_only=True, data_only=True)
    assert check["Giao dịch"].max_row == len(ROWS) + 1
    assert check["Tổng hợp ngày"].max_row == len(dates) + 2
    print(output.resolve())
    print(f"transactions={len(ROWS)}")
    print(f"income={sum(max(r[1], 0) for r in ROWS)}")
    print(f"expense={sum(abs(min(r[1], 0)) for r in ROWS)}")
    print(f"net={sum(r[1] for r in ROWS)}")


if __name__ == "__main__":
    main()
